import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook

def main():
    # Prompt for the input file path and strip any quotes
    file_path = input("Enter the full file path of your Excel file (e.g., C:\\path\\to\\file.xlsx): ").strip().strip('\'"')
    if not os.path.exists(file_path):
        print("The file does not exist. Exiting.")
        return

    # Load the Excel file and list available sheets
    try:
        xls = pd.ExcelFile(file_path)
        sheet_names = xls.sheet_names
        print("Available sheets in the workbook:")
        for idx, sheet in enumerate(sheet_names):
            print(f"  {idx+1}. {sheet}")
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    # Use the "parts" sheet if it exists; otherwise, prompt for a sheet name
    if "parts" in sheet_names:
        sheet_to_read = "parts"
    else:
        print("The sheet 'parts' was not found in the workbook.")
        chosen_sheet = input("Please enter the name of the sheet you want to process (or press Enter to exit): ").strip().strip('\'"')
        if not chosen_sheet or chosen_sheet not in sheet_names:
            print("Invalid or no sheet name entered. Exiting.")
            return
        sheet_to_read = chosen_sheet

    # Read the selected sheet into a DataFrame
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_to_read)
    except Exception as e:
        print(f"Error reading the sheet '{sheet_to_read}': {e}")
        return

    # Replace empty strings with NaN for the 'id' column
    if 'id' in df.columns:
        df['id'] = df['id'].replace('', np.nan)
    else:
        print("The 'id' column was not found in the sheet. Exiting.")
        return

    # Verify the existence of the 'Mfg Part Number' column
    if 'Mfg Part Number' not in df.columns:
        print("The 'Mfg Part Number' column was not found in the sheet. Exiting.")
        return

    # Save the original id values before processing
    df['original_id'] = df['id']

    # Define a helper function to fill missing 'id' values within each group
    def fill_missing_ids(s):
        non_missing = s.dropna()
        if not non_missing.empty:
            fill_val = non_missing.iloc[0]
            return s.fillna(fill_val)
        else:
            return s

    # Process the DataFrame: fill missing 'id' values for each Mfg Part Number group
    df['id'] = df.groupby('Mfg Part Number')['id'].transform(fill_missing_ids)

    # Create a DataFrame containing only duplicate parts based on "Mfg Part Number"
    duplicate_df = df[df.duplicated(subset=["Mfg Part Number"], keep=False)]

    # Create a DataFrame containing only rows where an ID was assigned
    # (i.e. originally missing but now filled)
    assigned_df = df[df['original_id'].isna() & df['id'].notna()]

    # Remove existing sheets if they exist
    try:
        book = load_workbook(file_path)
        sheets_to_remove = ["updated_parts_list", "duplicate_parts_found", "assigned_ids"]
        for sheet in sheets_to_remove:
            if sheet in book.sheetnames:
                ws = book[sheet]
                book.remove(ws)
        book.save(file_path)
    except Exception as e:
        print(f"Error removing existing sheets: {e}")

    # Write the new sheets into the workbook using append mode
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Write full processed data to "updated_parts_list"
            df.to_excel(writer, sheet_name="updated_parts_list", index=False)
            # Write duplicate rows to "duplicate_parts_found"
            duplicate_df.to_excel(writer, sheet_name="duplicate_parts_found", index=False)
            # Write rows that got assigned IDs to "assigned_ids"
            assigned_df.to_excel(writer, sheet_name="assigned_ids", index=False)
        print(f"'updated_parts_list', 'duplicate_parts_found', and 'assigned_ids' sheets have been added to {file_path}.")
    except Exception as e:
        print(f"Error saving updated sheets: {e}")

if __name__ == '__main__':
    main()
