import os
import time
import piexif
import logging
from PIL import Image
from openpyxl import load_workbook

# Predefined EXIF tags (as used in your load sheet)
exif_tags = [
    {"Tag Name": "ImageDescription", "Tag ID": "0x010E", "IFD": "0th", "Data Type": "ASCII"},
    {"Tag Name": "Make", "Tag ID": "0x010F", "IFD": "0th", "Data Type": "ASCII"},
    {"Tag Name": "Model", "Tag ID": "0x0110", "IFD": "0th", "Data Type": "ASCII"},
    {"Tag Name": "Orientation", "Tag ID": "0x0112", "IFD": "0th", "Data Type": "SHORT"},
    {"Tag Name": "XResolution", "Tag ID": "0x011A", "IFD": "0th", "Data Type": "RATIONAL"},
    {"Tag Name": "YResolution", "Tag ID": "0x011B", "IFD": "0th", "Data Type": "RATIONAL"},
    {"Tag Name": "ResolutionUnit", "Tag ID": "0x0128", "IFD": "0th", "Data Type": "SHORT"},
    {"Tag Name": "Software", "Tag ID": "0x0131", "IFD": "0th", "Data Type": "ASCII"},
    {"Tag Name": "DateTime", "Tag ID": "0x0132", "IFD": "0th", "Data Type": "ASCII"},
    {"Tag Name": "Artist", "Tag ID": "0x013B", "IFD": "0th", "Data Type": "ASCII"},
    {"Tag Name": "Copyright", "Tag ID": "0x8298", "IFD": "0th", "Data Type": "ASCII"},
    {"Tag Name": "ExposureTime", "Tag ID": "0x829A", "IFD": "Exif", "Data Type": "RATIONAL"},
    {"Tag Name": "FNumber", "Tag ID": "0x829D", "IFD": "Exif", "Data Type": "RATIONAL"},
    {"Tag Name": "ExposureProgram", "Tag ID": "0x8822", "IFD": "Exif", "Data Type": "SHORT"},
    {"Tag Name": "ISOSpeedRatings", "Tag ID": "0x8827", "IFD": "Exif", "Data Type": "SHORT"},
    {"Tag Name": "ExifVersion", "Tag ID": "0x9000", "IFD": "Exif", "Data Type": "UNDEFINED"},
    {"Tag Name": "DateTimeOriginal", "Tag ID": "0x9003", "IFD": "Exif", "Data Type": "ASCII"},
    {"Tag Name": "DateTimeDigitized", "Tag ID": "0x9004", "IFD": "Exif", "Data Type": "ASCII"},
    {"Tag Name": "ShutterSpeedValue", "Tag ID": "0x9201", "IFD": "Exif", "Data Type": "SRATIONAL"},
    {"Tag Name": "ApertureValue", "Tag ID": "0x9202", "IFD": "Exif", "Data Type": "RATIONAL"},
    {"Tag Name": "FocalLength", "Tag ID": "0x920A", "IFD": "Exif", "Data Type": "RATIONAL"},
    {"Tag Name": "UserComment", "Tag ID": "0x9286", "IFD": "Exif", "Data Type": "UNDEFINED"}
]

def parse_value(value, data_type):
    """
    Convert the cell value (a string) into the appropriate type based on the EXIF data type.
    For RATIONAL types, expect input in "numerator/denominator" format.
    """
    if value == "" or value is None:
        return None
    try:
        if data_type == "ASCII":
            return value  # piexif accepts Python strings for ASCII tags.
        elif data_type == "SHORT":
            return int(value)
        elif data_type in ["RATIONAL", "SRATIONAL"]:
            parts = value.split("/")
            if len(parts) == 2:
                numerator = int(parts[0].strip())
                denominator = int(parts[1].strip())
                return (numerator, denominator)
            else:
                num = int(value)
                return (num, 1)
        elif data_type == "UNDEFINED":
            return value
    except Exception as e:
        print(f"Error parsing value '{value}' as {data_type}: {e}")
        return None

# Setup logging to file
log_file = "exif_update_log.txt"
logging.basicConfig(filename=log_file, level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Prompt for the Excel load sheet file and the folder containing images.
excel_path = input("Enter the full path to the Excel load sheet file: ").strip()
image_folder = input("Enter the folder path containing the images: ").strip()

if not os.path.isfile(excel_path):
    print("The Excel load sheet file does not exist. Exiting.")
    exit(1)
if not os.path.isdir(image_folder):
    print("The image folder does not exist. Exiting.")
    exit(1)

# Load the Excel workbook and select the "Image Load Sheet" sheet.
wb = load_workbook(excel_path)
try:
    load_sheet = wb["Image Load Sheet"]
except KeyError:
    print("The workbook does not contain an 'Image Load Sheet'. Exiting.")
    exit(1)

# Process each row (skip header row).
# Expected columns:
#   Column 1: Original Image Name
#   Column 2: New Image Name (if you want to change the file name; otherwise blank)
#   Columns 3+: EXIF tag values (in the same order as defined in exif_tags)
for row in load_sheet.iter_rows(min_row=2, values_only=True):
    orig_image_name = row[0]
    new_image_name = row[1]
    if not orig_image_name:
        continue

    orig_image_path = os.path.join(image_folder, orig_image_name)
    if not os.path.isfile(orig_image_path):
        logging.warning(f"Image file {orig_image_path} not found. Skipping.")
        continue

    # Determine final image path.
    final_image_path = orig_image_path
    if new_image_name and new_image_name != orig_image_name:
        final_image_path = os.path.join(image_folder, new_image_name)

    logging.info(f"Processing image: {orig_image_name}")

    # --- Step 1: Open image to extract and update EXIF ---
    try:
        with Image.open(orig_image_path) as img:
            try:
                exif_dict = piexif.load(img.info.get("exif", b""))
            except Exception:
                exif_dict = {"0th": {}, "Exif": {}, "GPS": {}, "1st": {}, "thumbnail": None}

            for idx, tag in enumerate(exif_tags, start=2):
                cell_value = row[idx]
                parsed = parse_value(cell_value, tag["Data Type"])
                if parsed is not None:
                    ifd = tag["IFD"]
                    tag_id = int(tag["Tag ID"], 16)
                    current_value = exif_dict.get(ifd, {}).get(tag_id, None)
                    if isinstance(current_value, bytes):
                        try:
                            current_value = current_value.decode("utf-8", errors="replace")
                        except Exception:
                            current_value = str(current_value)
                    if current_value != parsed:
                        logging.info(f"{orig_image_name} - {tag['Tag Name']} changed from '{current_value}' to '{parsed}'")
                        exif_dict[ifd][tag_id] = parsed

            try:
                exif_bytes = piexif.dump(exif_dict)
            except Exception as e:
                logging.error(f"Failed to dump EXIF for {orig_image_name}: {e}")
                continue
    except Exception as e:
        logging.error(f"Failed to open {orig_image_path}: {e}")
        continue

    # --- Step 2: Wait briefly to ensure file is closed, then rename if needed ---
    time.sleep(0.1)
    if final_image_path != orig_image_path:
        try:
            os.rename(orig_image_path, final_image_path)
            logging.info(f"Renamed file from {orig_image_name} to {new_image_name}")
        except Exception as e:
            logging.error(f"Failed to rename {orig_image_name} to {new_image_name}: {e}")
            continue

    # --- Step 3: Reopen image and save updated EXIF data ---
    try:
        with Image.open(final_image_path) as img:
            img.save(final_image_path, exif=exif_bytes)
            logging.info(f"Successfully updated EXIF for {new_image_name if new_image_name else orig_image_name}")
    except Exception as e:
        logging.error(f"Failed to update EXIF for {new_image_name if new_image_name else orig_image_name}: {e}")

print("EXIF update process complete. Check the log file for details:", log_file)
