#!/usr/bin/env python3
"""
Amend Existing Excel with Academic Data

This script opens an existing Excel file, adds or updates worksheets with academic subject data,
and saves the amended file.

Usage:
    python AmendExistingExcelWithAcademicData.py
"""

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil

# File paths
SOURCE_FILE = r"C:\Users\10169062\Desktop\AU_IndusMaintdb\Database\DB_LOADSHEETS\load_equipment_relationships_table_data - Copy.xlsx"
BACKUP_FILE = SOURCE_FILE.replace(".xlsx", "_backup.xlsx")
TEMP_FILE = SOURCE_FILE.replace(".xlsx", "_temp.xlsx")


# Define area data
def create_area_data():
    """Create the Area data (top level)."""
    return [{
        "id": 1,  # Simulated ID
        "name": "Academic",
        "description": "Academic and technical knowledge across various fields"
    }]


# Define equipment group data (Academic Fields)
def create_equipment_group_data():
    """Create the Equipment Group data (academic fields)."""
    # Initialize with a simulated ID counter
    id_counter = 1

    equipment_groups = []

    # Business Administration
    equipment_groups.append({
        "id": id_counter,
        "name": "Business Administration",
        "description": "Study of management and operation of business enterprises",
        "area_id": 1  # Reference to Academic area
    })
    business_id = id_counter
    id_counter += 1

    # Physics
    equipment_groups.append({
        "id": id_counter,
        "name": "Physics",
        "description": "Study of matter, energy, and their interactions",
        "area_id": 1
    })
    physics_id = id_counter
    id_counter += 1

    # Chemistry
    equipment_groups.append({
        "id": id_counter,
        "name": "Chemistry",
        "description": "Study of substances, their properties, and reactions",
        "area_id": 1
    })
    chemistry_id = id_counter
    id_counter += 1

    # Biology
    equipment_groups.append({
        "id": id_counter,
        "name": "Biology",
        "description": "Study of living organisms",
        "area_id": 1
    })
    biology_id = id_counter
    id_counter += 1

    # Pure Mathematics
    equipment_groups.append({
        "id": id_counter,
        "name": "Pure Mathematics",
        "description": "Study of abstract concepts and structures",
        "area_id": 1
    })
    pure_math_id = id_counter
    id_counter += 1

    # Applied Mathematics
    equipment_groups.append({
        "id": id_counter,
        "name": "Applied Mathematics",
        "description": "Application of mathematical methods to solve real-world problems",
        "area_id": 1
    })
    applied_math_id = id_counter
    id_counter += 1

    # Computer Science
    equipment_groups.append({
        "id": id_counter,
        "name": "Computer Science",
        "description": "Study of computation, algorithms, and information processing",
        "area_id": 1
    })
    cs_id = id_counter
    id_counter += 1

    # Philosophy
    equipment_groups.append({
        "id": id_counter,
        "name": "Philosophy",
        "description": "Study of fundamental questions about existence, knowledge, ethics, and more",
        "area_id": 1
    })
    philosophy_id = id_counter
    id_counter += 1

    # History
    equipment_groups.append({
        "id": id_counter,
        "name": "History",
        "description": "Study of past events, societies, and civilizations",
        "area_id": 1
    })
    history_id = id_counter
    id_counter += 1

    # Literature
    equipment_groups.append({
        "id": id_counter,
        "name": "Literature",
        "description": "Study of written works of art",
        "area_id": 1
    })
    literature_id = id_counter
    id_counter += 1

    # Psychology
    equipment_groups.append({
        "id": id_counter,
        "name": "Psychology",
        "description": "Study of mind and behavior",
        "area_id": 1
    })
    psychology_id = id_counter
    id_counter += 1

    # Economics
    equipment_groups.append({
        "id": id_counter,
        "name": "Economics",
        "description": "Study of production, distribution, and consumption of goods and services",
        "area_id": 1
    })
    economics_id = id_counter
    id_counter += 1

    # Sociology
    equipment_groups.append({
        "id": id_counter,
        "name": "Sociology",
        "description": "Study of society, social relationships, and culture",
        "area_id": 1
    })
    sociology_id = id_counter
    id_counter += 1

    # Welding Technology
    equipment_groups.append({
        "id": id_counter,
        "name": "Welding Technology",
        "description": "Study of joining materials through fusion processes",
        "area_id": 1
    })
    welding_id = id_counter
    id_counter += 1

    # Machining Technology
    equipment_groups.append({
        "id": id_counter,
        "name": "Machining Technology",
        "description": "Study of material removal processes to create parts",
        "area_id": 1
    })
    machining_id = id_counter
    id_counter += 1

    # Industrial Electrical Systems
    equipment_groups.append({
        "id": id_counter,
        "name": "Industrial Electrical Systems",
        "description": "Study of electrical systems in industrial settings",
        "area_id": 1
    })
    electrical_id = id_counter
    id_counter += 1

    # Fluid Power Systems
    equipment_groups.append({
        "id": id_counter,
        "name": "Fluid Power Systems",
        "description": "Study of hydraulic and pneumatic systems for power transmission",
        "area_id": 1
    })
    fluid_power_id = id_counter
    id_counter += 1

    # Store the IDs for use in creating model data
    global equipment_group_ids
    equipment_group_ids = {
        "business": business_id,
        "physics": physics_id,
        "chemistry": chemistry_id,
        "biology": biology_id,
        "pure_math": pure_math_id,
        "applied_math": applied_math_id,
        "cs": cs_id,
        "philosophy": philosophy_id,
        "history": history_id,
        "literature": literature_id,
        "psychology": psychology_id,
        "economics": economics_id,
        "sociology": sociology_id,
        "welding": welding_id,
        "machining": machining_id,
        "electrical": electrical_id,
        "fluid_power": fluid_power_id
    }

    return equipment_groups


# Define model data (Subjects)
def create_model_data():
    """Create the Model data (subjects/branches)."""
    models = []
    id_counter = 1

    # Business Administration subjects
    business_subjects = [
        {"name": "Marketing", "description": "Study of promoting and selling products or services"},
        {"name": "Finance", "description": "Study of money management and asset allocation"},
        {"name": "Management", "description": "Study of organizational leadership and administration"},
        {"name": "Accounting",
         "description": "Study of recording, classifying, and summarizing financial transactions"},
        {"name": "Human Resources", "description": "Study of managing an organization's workforce"},
        {"name": "Operations Management", "description": "Study of designing and controlling production processes"},
        {"name": "Business Analytics", "description": "Study of data analysis techniques for business insights"},
        {"name": "Entrepreneurship", "description": "Study of starting and running new business ventures"},
        {"name": "International Business", "description": "Study of global business operations and strategies"}
    ]

    for subject in business_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["business"]
        })
        id_counter += 1

    # Physics subjects
    physics_subjects = [
        {"name": "Mechanics", "description": "Study of motion and forces"},
        {"name": "Quantum Physics", "description": "Study of subatomic particles and their behaviors"},
        {"name": "Thermodynamics", "description": "Study of heat, energy, and work"},
        {"name": "Electromagnetism", "description": "Study of electromagnetic force and fields"},
        {"name": "Relativity", "description": "Study of space, time, and gravity"}
    ]

    for subject in physics_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["physics"]
        })
        id_counter += 1

    # Chemistry subjects
    chemistry_subjects = [
        {"name": "Organic Chemistry", "description": "Study of carbon-containing compounds"},
        {"name": "Inorganic Chemistry", "description": "Study of non-carbon compounds"},
        {"name": "Physical Chemistry", "description": "Study of how matter behaves on a molecular and atomic level"},
        {"name": "Biochemistry", "description": "Study of chemical processes within living organisms"},
        {"name": "Analytical Chemistry",
         "description": "Study of separation, identification, and quantification of matter"}
    ]

    for subject in chemistry_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["chemistry"]
        })
        id_counter += 1

    # Biology subjects
    biology_subjects = [
        {"name": "Molecular Biology", "description": "Study of biological activity at the molecular level"},
        {"name": "Genetics", "description": "Study of genes, heredity, and genetic variation"},
        {"name": "Ecology", "description": "Study of interactions between organisms and their environment"}
    ]

    for subject in biology_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["biology"]
        })
        id_counter += 1

    # Pure Mathematics subjects
    pure_math_subjects = [
        {"name": "Algebra", "description": "Study of mathematical symbols and rules"},
        {"name": "Calculus", "description": "Study of continuous change and functions"},
        {"name": "Geometry", "description": "Study of shapes, sizes, and properties of space"},
        {"name": "Number Theory", "description": "Study of integers and integer-valued functions"},
        {"name": "Topology", "description": "Study of properties preserved under continuous deformations"}
    ]

    for subject in pure_math_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["pure_math"]
        })
        id_counter += 1

    # Applied Mathematics subjects
    applied_math_subjects = [
        {"name": "Statistics", "description": "Study of data collection, analysis, and interpretation"},
        {"name": "Operations Research", "description": "Application of analytical methods for decision-making"},
        {"name": "Mathematical Modeling", "description": "Using mathematics to describe real-world phenomena"},
        {"name": "Financial Mathematics", "description": "Application of mathematical methods in finance"}
    ]

    for subject in applied_math_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["applied_math"]
        })
        id_counter += 1

    # Computer Science subjects
    cs_subjects = [
        {"name": "Algorithms", "description": "Study of computational procedures and problem-solving methods"},
        {"name": "Data Structures", "description": "Study of organizing and storing data efficiently"},
        {"name": "Artificial Intelligence",
         "description": "Study of intelligent agent development and machine learning"},
        {"name": "Database Systems", "description": "Study of data organization, storage, and retrieval methods"},
        {"name": "Computer Networks", "description": "Study of data communication systems and protocols"},
        {"name": "Software Engineering", "description": "Study of systematic development of software applications"},
        {"name": "Cybersecurity",
         "description": "Study of protecting computer systems from unauthorized access and attacks"},
        {"name": "Theoretical Computer Science", "description": "Mathematical study of computation and algorithms"},
        {"name": "Human-Computer Interaction", "description": "Study of interaction between humans and computers"}
    ]

    for subject in cs_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["cs"]
        })
        id_counter += 1

    # Philosophy subjects
    philosophy_subjects = [
        {"name": "Ethics", "description": "Study of moral principles and values"},
        {"name": "Epistemology", "description": "Study of knowledge, belief, and justification"},
        {"name": "Metaphysics", "description": "Study of reality, existence, and being"},
        {"name": "Logic", "description": "Study of valid reasoning and inference"}
    ]

    for subject in philosophy_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["philosophy"]
        })
        id_counter += 1

    # History subjects
    history_subjects = [
        {"name": "World History", "description": "Study of history on a global scale"},
        {"name": "U.S. History", "description": "Study of United States history"},
        {"name": "Ancient History", "description": "Study of early human civilizations"},
        {"name": "Medieval History", "description": "Study of the Middle Ages"},
        {"name": "Modern History", "description": "Study of recent centuries of human history"}
    ]

    for subject in history_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["history"]
        })
        id_counter += 1

    # Literature subjects
    literature_subjects = [
        {"name": "English Literature", "description": "Study of literature written in English"},
        {"name": "World Literature", "description": "Study of literature from various cultures and languages"},
        {"name": "Poetry", "description": "Study of poetic works and forms"},
        {"name": "Drama", "description": "Study of theatrical works"}
    ]

    for subject in literature_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["literature"]
        })
        id_counter += 1

    # Psychology subjects
    psychology_subjects = [
        {"name": "Clinical Psychology", "description": "Study and treatment of mental illness and distress"},
        {"name": "Cognitive Psychology",
         "description": "Study of mental processes including perception, thinking, and memory"},
        {"name": "Developmental Psychology", "description": "Study of psychological growth across the lifespan"},
        {"name": "Social Psychology",
         "description": "Study of how individuals' thoughts and behavior are influenced by others"}
    ]

    for subject in psychology_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["psychology"]
        })
        id_counter += 1

    # Economics subjects
    economics_subjects = [
        {"name": "Microeconomics",
         "description": "Study of individual and business decisions regarding resource allocation"},
        {"name": "Macroeconomics",
         "description": "Study of economy-wide phenomena like inflation, growth, and unemployment"},
        {"name": "International Economics", "description": "Study of economic interactions between countries"},
        {"name": "Econometrics", "description": "Application of statistical methods to economic data"}
    ]

    for subject in economics_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["economics"]
        })
        id_counter += 1

    # Sociology subjects
    sociology_subjects = [
        {"name": "Cultural Sociology", "description": "Study of the influence of culture on social life"},
        {"name": "Urban Sociology", "description": "Study of social life and interactions in urban environments"},
        {"name": "Social Inequality", "description": "Study of social differences and hierarchies"}
    ]

    for subject in sociology_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["sociology"]
        })
        id_counter += 1

    # Welding Technology subjects
    welding_subjects = [
        {"name": "Arc Welding", "description": "Welding processes using an electric arc (SMAW, GMAW, GTAW)"},
        {"name": "Resistance Welding",
         "description": "Welding processes that use electrical resistance to generate heat"},
        {"name": "Oxyfuel Welding", "description": "Welding using fuel gases and oxygen to produce a flame"},
        {"name": "Welding Metallurgy", "description": "Study of metal properties and behaviors during welding"},
        {"name": "Welding Inspection", "description": "Quality control and testing of welded joints"},
        {"name": "Welding Automation", "description": "Automated and robotic welding systems and programming"}
    ]

    for subject in welding_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["welding"]
        })
        id_counter += 1

    # Machining Technology subjects
    machining_subjects = [
        {"name": "CNC Machining", "description": "Computer numerical control machining processes and programming"},
        {"name": "Manual Machining",
         "description": "Traditional machine tool operation (lathes, mills, drill presses)"},
        {"name": "Precision Measurement", "description": "Metrology techniques and instruments for machined parts"},
        {"name": "CAD/CAM Systems", "description": "Computer-aided design and manufacturing for machining"},
        {"name": "Advanced Machining Processes",
         "description": "Non-traditional processes like EDM, waterjet, and laser cutting"},
        {"name": "Tool Design", "description": "Design of cutting tools, fixtures, and machine tooling"}
    ]

    for subject in machining_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["machining"]
        })
        id_counter += 1

    # Industrial Electrical Systems subjects
    electrical_subjects = [
        {"name": "Power Distribution", "description": "Industrial power systems and distribution networks"},
        {"name": "Motor Controls", "description": "Electric motor operation, control, and protection systems"},
        {"name": "Industrial Controls", "description": "PLCs, SCADA, and other industrial control systems"},
        {"name": "Electrical Troubleshooting", "description": "Diagnosing and repairing electrical system faults"},
        {"name": "Electrical Safety", "description": "Hazard identification and safety practices for electrical work"},
        {"name": "Industrial IoT", "description": "Internet of Things applications in industrial settings"}
    ]

    for subject in electrical_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["electrical"]
        })
        id_counter += 1

    # Fluid Power Systems subjects
    fluid_power_subjects = [
        {"name": "Hydraulic Systems", "description": "Liquid-based power transmission systems design and operation"},
        {"name": "Pneumatic Systems", "description": "Compressed air power systems design and operation"},
        {"name": "Fluid Power Components",
         "description": "Pumps, valves, actuators, and other hydraulic/pneumatic components"},
        {"name": "Fluid Power Maintenance", "description": "Troubleshooting, repair, and preventive maintenance"},
        {"name": "Electrohydraulic Systems", "description": "Integration of electronics with hydraulic systems"},
        {"name": "Fluid Power Circuit Design",
         "description": "Designing and analyzing hydraulic and pneumatic circuits"}
    ]

    for subject in fluid_power_subjects:
        models.append({
            "id": id_counter,
            "name": subject["name"],
            "description": subject["description"],
            "equipment_group_id": equipment_group_ids["fluid_power"]
        })
        id_counter += 1

    return models


# Create relationship data
def create_relationship_data(areas, equipment_groups, models):
    """Create a table showing relationships between entities."""
    relationships = []

    # For each equipment group, add its relationship with its parent area
    for group in equipment_groups:
        relationships.append({
            "level_1_type": "Area",
            "level_1_id": group["area_id"],
            "level_1_name": next(area["name"] for area in areas if area["id"] == group["area_id"]),
            "level_2_type": "EquipmentGroup",
            "level_2_id": group["id"],
            "level_2_name": group["name"],
            "level_3_type": None,
            "level_3_id": None,
            "level_3_name": None
        })

    # For each model, add its relationship with its parent equipment group and grandparent area
    for model in models:
        parent_group = next(group for group in equipment_groups if group["id"] == model["equipment_group_id"])
        relationships.append({
            "level_1_type": "Area",
            "level_1_id": parent_group["area_id"],
            "level_1_name": next(area["name"] for area in areas if area["id"] == parent_group["area_id"]),
            "level_2_type": "EquipmentGroup",
            "level_2_id": parent_group["id"],
            "level_2_name": parent_group["name"],
            "level_3_type": "Model",
            "level_3_id": model["id"],
            "level_3_name": model["name"]
        })

    return relationships


# Format worksheets with styling
def format_worksheet(worksheet, df):
    """Apply formatting to worksheet."""
    # Set column widths
    for i, column in enumerate(df.columns):
        worksheet.column_dimensions[get_column_letter(i + 1)].width = max(15, len(column) + 2)

    # Format header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, cell in enumerate(worksheet[1]):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add borders to all cells
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # Not header
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Freeze top row
    worksheet.freeze_panes = 'A2'


def create_backup():
    """Create a backup of the source file."""
    try:
        print(f"Creating backup of {SOURCE_FILE}...")
        shutil.copy2(SOURCE_FILE, BACKUP_FILE)
        print(f"Backup created at {BACKUP_FILE}")
        return True
    except Exception as e:
        print(f"Error creating backup: {str(e)}")
        return False


def check_source_file_exists():
    """Check if the source file exists."""
    if not os.path.exists(SOURCE_FILE):
        print(f"Error: Source file not found at {SOURCE_FILE}")
        return False
    return True


def amend_excel_file():
    """Amend the existing Excel file with academic data."""
    # Check if source file exists
    if not check_source_file_exists():
        return False

    # Create backup
    if not create_backup():
        return False

    try:
        print("Generating academic data...")
        # Create the data
        areas = create_area_data()
        equipment_groups = create_equipment_group_data()
        models = create_model_data()
        relationships = create_relationship_data(areas, equipment_groups, models)

        # Convert to pandas DataFrames
        areas_df = pd.DataFrame(areas)
        equipment_groups_df = pd.DataFrame(equipment_groups)
        models_df = pd.DataFrame(models)
        relationships_df = pd.DataFrame(relationships)

        print(f"Loading existing workbook from {SOURCE_FILE}...")
        try:
            # Try to load the existing workbook
            wb = load_workbook(SOURCE_FILE)
            print("Existing workbook loaded successfully")
        except Exception as e:
            print(f"Error loading existing workbook: {str(e)}")
            print("Creating new workbook instead...")
            wb = Workbook()
            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)

        # Add or replace Areas sheet
        if "Areas" in wb.sheetnames:
            print("Replacing existing Areas sheet...")
            # Remove the existing sheet
            del wb["Areas"]

        print("Adding Areas sheet...")
        areas_sheet = wb.create_sheet("Areas")
        for r_idx, row in enumerate(areas_df.itertuples(index=False), 1):
            if r_idx == 1:
                # Write header
                for c_idx, col_name in enumerate(areas_df.columns, 1):
                    areas_sheet.cell(row=r_idx, column=c_idx).value = col_name
            # Write data
            for c_idx, value in enumerate(row, 1):
                areas_sheet.cell(row=r_idx + 1, column=c_idx).value = value

        # Add or replace Equipment Groups sheet
        if "EquipmentGroups" in wb.sheetnames:
            print("Replacing existing EquipmentGroups sheet...")
            del wb["EquipmentGroups"]

        print("Adding EquipmentGroups sheet...")
        eg_sheet = wb.create_sheet("EquipmentGroups")
        for r_idx, row in enumerate(equipment_groups_df.itertuples(index=False), 1):
            if r_idx == 1:
                # Write header
                for c_idx, col_name in enumerate(equipment_groups_df.columns, 1):
                    eg_sheet.cell(row=r_idx, column=c_idx).value = col_name
            # Write data
            for c_idx, value in enumerate(row, 1):
                eg_sheet.cell(row=r_idx + 1, column=c_idx).value = value

        # Add or replace Models sheet
        if "Models" in wb.sheetnames:
            print("Replacing existing Models sheet...")
            del wb["Models"]

        print("Adding Models sheet...")
        models_sheet = wb.create_sheet("Models")
        for r_idx, row in enumerate(models_df.itertuples(index=False), 1):
            if r_idx == 1:
                # Write header
                for c_idx, col_name in enumerate(models_df.columns, 1):
                    models_sheet.cell(row=r_idx, column=c_idx).value = col_name
            # Write data
            for c_idx, value in enumerate(row, 1):
                models_sheet.cell(row=r_idx + 1, column=c_idx).value = value

        # Add or replace Relationships sheet
        if "Relationships" in wb.sheetnames:
            print("Replacing existing Relationships sheet...")
            del wb["Relationships"]

        print("Adding Relationships sheet...")
        rel_sheet = wb.create_sheet("Relationships")
        for r_idx, row in enumerate(relationships_df.itertuples(index=False), 1):
            if r_idx == 1:
                # Write header
                for c_idx, col_name in enumerate(relationships_df.columns, 1):
                    rel_sheet.cell(row=r_idx, column=c_idx).value = col_name
            # Write data
            for c_idx, value in enumerate(row, 1):
                rel_sheet.cell(row=r_idx + 1, column=c_idx).value = value

        # Add or replace Hierarchy View sheet
        if "Hierarchy View" in wb.sheetnames:
            print("Replacing existing Hierarchy View sheet...")
            del wb["Hierarchy View"]

        print("Adding Hierarchy View sheet...")
        hierarchy_sheet = wb.create_sheet("Hierarchy View")
        hierarchy_sheet.append(["Level", "ID", "Name", "Description", "Parent"])

        # Add Area (level 1)
        row_index = 2
        for area in areas:
            hierarchy_sheet.cell(row=row_index, column=1).value = "Area"
            hierarchy_sheet.cell(row=row_index, column=2).value = area["id"]
            hierarchy_sheet.cell(row=row_index, column=3).value = area["name"]
            hierarchy_sheet.cell(row=row_index, column=4).value = area["description"]
            hierarchy_sheet.cell(row=row_index, column=5).value = "None"

            # Apply formatting
            for col in range(1, 6):
                cell = hierarchy_sheet.cell(row=row_index, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

            row_index += 1

            # Add Equipment Groups (level 2) for this area
            for group in equipment_groups:
                if group["area_id"] == area["id"]:
                    hierarchy_sheet.cell(row=row_index, column=1).value = "EquipmentGroup"
                    hierarchy_sheet.cell(row=row_index, column=2).value = group["id"]
                    hierarchy_sheet.cell(row=row_index, column=3).value = group["name"]
                    hierarchy_sheet.cell(row=row_index, column=4).value = group["description"]
                    hierarchy_sheet.cell(row=row_index, column=5).value = f"Area: {area['name']}"

                    # Apply formatting
                    for col in range(1, 6):
                        cell = hierarchy_sheet.cell(row=row_index, column=col)
                        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

                    group_row = row_index
                    row_index += 1

                    # Add Models (level 3) for this equipment group
                    for model in models:
                        if model["equipment_group_id"] == group["id"]:
                            hierarchy_sheet.cell(row=row_index, column=1).value = "Model"
                            hierarchy_sheet.cell(row=row_index, column=2).value = model["id"]
                            hierarchy_sheet.cell(row=row_index, column=3).value = model["name"]
                            hierarchy_sheet.cell(row=row_index, column=4).value = model["description"]
                            hierarchy_sheet.cell(row=row_index, column=5).value = f"EquipmentGroup: {group['name']}"
                            row_index += 1

        # Format all worksheets
        print("Applying formatting to all worksheets...")
        format_worksheet(areas_sheet, areas_df)
        format_worksheet(eg_sheet, equipment_groups_df)
        format_worksheet(models_sheet, models_df)
        format_worksheet(rel_sheet, relationships_df)

        # Format hierarchy sheet
        for column in ["A", "B", "C", "D", "E"]:
            hierarchy_sheet.column_dimensions[column].width = 30 if column in ["C", "D", "E"] else 15

        # Format header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for cell in hierarchy_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Add borders to all cells
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in hierarchy_sheet.iter_rows(min_row=1, max_row=hierarchy_sheet.max_row,
                                             min_col=1, max_col=hierarchy_sheet.max_column):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Not header
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Freeze top row
        hierarchy_sheet.freeze_panes = 'A2'

        # Set Hierarchy View as the active sheet
        wb.active = hierarchy_sheet

        # Save to a temporary file first
        print(f"Saving to temporary file {TEMP_FILE}...")
        wb.save(TEMP_FILE)

        # If the save was successful, replace the original file
        print(f"Replacing original file {SOURCE_FILE}...")
        shutil.move(TEMP_FILE, SOURCE_FILE)

        print(f"Excel file successfully amended at: {SOURCE_FILE}")
        print("Summary of data added:")
        print(f"- Areas: {len(areas)}")
        print(f"- Equipment Groups: {len(equipment_groups)}")
        print(f"- Models: {len(models)}")
        print(f"- Relationships: {len(relationships)}")

        return True
    except Exception as e:
        print(f"Error amending Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Main function to amend the Excel file."""
    print("Starting Amend Existing Excel with Academic Data script...")

    success = amend_excel_file()

    if success:
        print("Script completed successfully!")
    else:
        print("Script encountered errors. Check the output above for details.")


if __name__ == "__main__":
    main()