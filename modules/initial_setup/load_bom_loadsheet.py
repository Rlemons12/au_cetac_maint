import os
import sys
from openpyxl import load_workbook, Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import pandas as pd

# Import the new PostgreSQL framework components
from modules.configuration.config import DB_LOADSHEET_BOMS
from modules.initial_setup.initializer_logger import initializer_logger, close_initializer_logger
from modules.configuration.config_env import DatabaseConfig
from modules.configuration.log_config import (
    debug_id, info_id, warning_id, error_id,
    set_request_id, get_request_id, log_timed_operation
)


class PostgreSQLBOMLoadsheetProcessor:
    """PostgreSQL-enhanced BOM loadsheet processor with improved error handling and user experience."""

    def __init__(self):
        self.request_id = set_request_id()
        self.db_config = DatabaseConfig()
        info_id("Initialized PostgreSQL BOM Loadsheet Processor", self.request_id)

        # Statistics tracking
        self.stats = {
            'files_processed': 0,
            'sheets_created': 0,
            'matches_found': 0,
            'entries_added': 0,
            'errors_encountered': 0
        }

    def validate_source_file(self, file_path):
        """Validate the source file with enhanced checks."""
        info_id(f"Validating source file: {file_path}", self.request_id)

        # Remove quotes and normalize path
        file_path = file_path.strip('\"\'').strip()

        # Check if file exists
        if not os.path.exists(file_path):
            error_id(f"Source file does not exist: {file_path}", self.request_id)
            return False, "File does not exist"

        # Check if filename starts with 'bom_for_'
        filename = os.path.basename(file_path)
        if not filename.startswith("bom_for_"):
            error_id(f"File does not start with 'bom_for_': {filename}", self.request_id)
            return False, "Filename must start with 'bom_for_'"

        # Check if it's an Excel file
        if not filename.endswith(('.xlsx', '.xls')):
            error_id(f"File is not an Excel file: {filename}", self.request_id)
            return False, "File must be an Excel file (.xlsx or .xls)"

        # Check if file is readable
        try:
            wb = load_workbook(file_path, read_only=True)
            if "BOM" not in wb.sheetnames:
                error_id(f"File does not contain 'BOM' sheet: {filename}", self.request_id)
                return False, "File must contain a 'BOM' sheet"
            wb.close()
        except Exception as e:
            error_id(f"Error reading Excel file: {str(e)}", self.request_id)
            return False, f"Cannot read Excel file: {str(e)}"

        info_id(f"Source file validation successful: {filename}", self.request_id)
        return True, "Valid"

    def prompt_for_source_file(self):
        """Enhanced source file selection with validation."""
        print("\n📂 BOM Source File Selection")
        print("=" * 40)
        print("Please provide the path to your BOM source file.")
        print("💡 Requirements:")
        print("   • Filename must start with 'bom_for_'")
        print("   • Must be an Excel file (.xlsx or .xls)")
        print("   • Must contain a 'BOM' sheet")
        print()

        max_attempts = 3
        attempt = 0

        while attempt < max_attempts:
            attempt += 1
            file_path = input(f"📂 Enter file path (attempt {attempt}/{max_attempts}): ").strip()

            if not file_path:
                print("⚠️  Please enter a valid file path")
                continue

            is_valid, message = self.validate_source_file(file_path)

            if is_valid:
                print(f"✅ Source file validated: {os.path.basename(file_path)}")
                info_id(f"Source file selected: {file_path}", self.request_id)
                return file_path
            else:
                print(f"❌ Validation failed: {message}")
                if attempt < max_attempts:
                    print("Please try again.")

        error_id("Failed to select valid source file after maximum attempts", self.request_id)
        raise ValueError("Unable to select a valid source file")

    def generate_target_file_path(self, source_file):
        """Generate target file path with enhanced directory handling."""
        info_id("Generating target file path", self.request_id)

        # Extract suffix from source filename
        source_filename = os.path.basename(source_file)
        suffix = source_filename.replace("bom_for_", "").replace(".xlsx", "").replace(".xls", "")

        # Ensure target directory exists
        try:
            if not os.path.exists(DB_LOADSHEET_BOMS):
                os.makedirs(DB_LOADSHEET_BOMS, exist_ok=True)
                info_id(f"Created target directory: {DB_LOADSHEET_BOMS}", self.request_id)
                print(f"📁 Created directory: {DB_LOADSHEET_BOMS}")
        except Exception as e:
            error_id(f"Error creating target directory: {str(e)}", self.request_id)
            raise

        # Generate target filename
        target_filename = f"load_bom_for_{suffix}.xlsx"
        target_path = os.path.join(DB_LOADSHEET_BOMS, target_filename)

        # Check if target file exists
        if os.path.exists(target_path):
            info_id(f"Target file already exists: {target_path}", self.request_id)
            print(f"📋 Target file exists and will be updated: {target_filename}")
        else:
            info_id(f"Creating new target file: {target_path}", self.request_id)
            print(f"📄 Creating new target file: {target_filename}")

        return target_path

    def copy_bom_sheet_with_validation(self, source_path, target_path):
        """Copy BOM sheet with enhanced error handling and validation."""
        info_id(f"Copying BOM sheet from {source_path} to {target_path}", self.request_id)
        print(f"📋 Processing BOM sheet...")

        try:
            with log_timed_operation("copy_bom_sheet", self.request_id):
                # Load source workbook
                wb_source = load_workbook(source_path)

                if "BOM" not in wb_source.sheetnames:
                    raise ValueError("Source workbook does not contain 'BOM' sheet")

                bom_sheet = wb_source["BOM"]

                # Count rows in source BOM
                source_row_count = bom_sheet.max_row
                info_id(f"Source BOM sheet contains {source_row_count} rows", self.request_id)

                # Load or create target workbook
                if os.path.exists(target_path):
                    wb_target = load_workbook(target_path)
                    info_id(f"Opened existing target workbook", self.request_id)
                else:
                    wb_target = Workbook()
                    # Remove default sheet
                    if 'Sheet' in wb_target.sheetnames:
                        wb_target.remove(wb_target['Sheet'])
                    info_id(f"Created new target workbook", self.request_id)

                # Generate new BOM sheet name
                suffix = os.path.basename(target_path).replace("load_bom_for_", "").replace(".xlsx", "")
                new_bom_sheet_name = f"bom_{suffix}"

                # Remove existing BOM sheet if it exists
                if new_bom_sheet_name in wb_target.sheetnames:
                    wb_target.remove(wb_target[new_bom_sheet_name])
                    warning_id(f"Removed existing sheet: {new_bom_sheet_name}", self.request_id)

                # Create new BOM sheet
                bom_target_sheet = wb_target.create_sheet(new_bom_sheet_name)
                info_id(f"Created new sheet: {new_bom_sheet_name}", self.request_id)

                # Copy data with progress tracking
                rows_copied = 0
                for row in bom_sheet.iter_rows(values_only=True):
                    bom_target_sheet.append(row)
                    rows_copied += 1

                info_id(f"Copied {rows_copied} rows to {new_bom_sheet_name}", self.request_id)
                print(f"   ✅ Copied {rows_copied} rows")

                # Create part_position_image sheet if it doesn't exist
                if "part_position_image" not in wb_target.sheetnames:
                    part_position_image_sheet = wb_target.create_sheet("part_position_image")
                    part_position_image_sheet.append(["part", "position", "image", "description"])
                    info_id("Created part_position_image sheet", self.request_id)
                    print(f"   📊 Created part_position_image sheet")
                    self.stats['sheets_created'] += 1

                # Save target workbook
                wb_target.save(target_path)
                info_id(f"Saved target workbook: {target_path}", self.request_id)
                print(f"   💾 Saved target workbook")

                self.stats['files_processed'] += 1

        except Exception as e:
            error_id(f"Error copying BOM sheet: {str(e)}", self.request_id, exc_info=True)
            self.stats['errors_encountered'] += 1
            raise

    def process_single_match(self, part_position_image_sheet, item_number, photo, description,
                             manufacturer_description):
        """Process a single match with enhanced error handling."""
        try:
            if photo and str(photo).strip():
                full_description = f"{description}, {manufacturer_description}" if manufacturer_description else description
                part_position_image_sheet.append([item_number, "", photo, full_description])
                debug_id(f"Added entry: {item_number} -> {photo}", self.request_id)
                self.stats['entries_added'] += 1
                return True
            return False
        except Exception as e:
            error_id(f"Error processing match for {item_number}: {str(e)}", self.request_id)
            self.stats['errors_encountered'] += 1
            return False

    def match_items_and_update_sheet(self, target_path, max_matches=5):
        """Enhanced item matching with progress tracking and configurability."""
        info_id("Starting item matching and sheet update process", self.request_id)
        print(f"🔍 Matching items and updating sheet...")

        try:
            with log_timed_operation("match_items_update_sheet", self.request_id):
                # Load target workbook
                wb_target = load_workbook(target_path)

                # Find BOM sheet
                bom_sheet_name = None
                for sheet_name in wb_target.sheetnames:
                    if sheet_name.startswith("bom_"):
                        bom_sheet_name = sheet_name
                        break

                if not bom_sheet_name:
                    raise ValueError("No BOM sheet found in target workbook")

                bom_sheet = wb_target[bom_sheet_name]
                part_position_image_sheet = wb_target["part_position_image"]

                # Load part list image file
                part_list_image_path = os.path.join(DB_LOADSHEET_BOMS, "part_list_image.xlsx")

                if not os.path.exists(part_list_image_path):
                    warning_id(f"Part list image file not found: {part_list_image_path}", self.request_id)
                    print(f"⚠️  Part list image file not found: {os.path.basename(part_list_image_path)}")
                    print(f"   Skipping item matching step")
                    return

                wb_part_list = load_workbook(part_list_image_path)

                if "photo_list" not in wb_part_list.sheetnames:
                    warning_id("'photo_list' sheet not found in part list image file", self.request_id)
                    print(f"⚠️  'photo_list' sheet not found in part list image file")
                    return

                photo_list_sheet = wb_part_list["photo_list"]

                # Log sheet headers for debugging
                bom_headers = [cell.value for cell in bom_sheet[1]]
                photo_list_headers = [cell.value for cell in photo_list_sheet[1]]
                debug_id(f"BOM headers: {bom_headers}", self.request_id)
                debug_id(f"Photo list headers: {photo_list_headers}", self.request_id)

                # Process matching with progress tracking
                match_count = 0
                processed_items = 0

                print(f"   🔄 Processing items (max {max_matches} matches)...")

                # Use ThreadPoolExecutor for parallel processing
                with ThreadPoolExecutor(max_workers=4) as executor:
                    futures = []

                    for row in bom_sheet.iter_rows(min_row=2, values_only=True):
                        if match_count >= max_matches:
                            break

                        processed_items += 1
                        item_number = str(row[3]) if len(row) > 3 and row[3] else None

                        if not item_number:
                            continue

                        # Remove leading 'A' if present
                        if item_number.startswith("A"):
                            item_number = item_number[1:]

                        part_number_prefix = item_number[:6] if len(item_number) >= 6 else item_number

                        # Search for matches in photo list
                        for photo_row in photo_list_sheet.iter_rows(min_row=2, values_only=True):
                            if match_count >= max_matches:
                                break

                            if not photo_row or not photo_row[0]:
                                continue

                            photo_part_number_prefix = str(photo_row[0])[:6]

                            if part_number_prefix == photo_part_number_prefix:
                                info_id(f"Match found: {item_number} -> {photo_row[0]}", self.request_id)
                                match_count += 1
                                self.stats['matches_found'] += 1

                                # Extract photo data
                                photo_a = photo_row[1] if len(photo_row) > 1 else None
                                desc_a = photo_row[4] if len(photo_row) > 4 else None
                                photo_b = photo_row[2] if len(photo_row) > 2 else None
                                desc_b = photo_row[5] if len(photo_row) > 5 else None
                                photo_c = photo_row[3] if len(photo_row) > 3 else None
                                desc_c = photo_row[6] if len(photo_row) > 6 else None
                                manufacturer_desc = photo_row[7] if len(photo_row) > 7 else None

                                # Submit processing tasks
                                futures.append(
                                    executor.submit(self.process_single_match, part_position_image_sheet,
                                                    item_number, photo_a, desc_a, manufacturer_desc)
                                )
                                futures.append(
                                    executor.submit(self.process_single_match, part_position_image_sheet,
                                                    item_number, photo_b, desc_b, manufacturer_desc)
                                )
                                futures.append(
                                    executor.submit(self.process_single_match, part_position_image_sheet,
                                                    item_number, photo_c, desc_c, manufacturer_desc)
                                )

                                break

                    # Wait for all tasks to complete
                    for future in as_completed(futures):
                        try:
                            future.result()
                        except Exception as e:
                            error_id(f"Error in parallel processing: {str(e)}", self.request_id)

                # Save updated workbook
                wb_target.save(target_path)

                info_id(f"Item matching completed. Processed {processed_items} items, found {match_count} matches",
                        self.request_id)
                print(f"   ✅ Found {match_count} matches from {processed_items} items")
                print(f"   💾 Updated part_position_image sheet")

        except Exception as e:
            error_id(f"Error in item matching: {str(e)}", self.request_id, exc_info=True)
            self.stats['errors_encountered'] += 1
            raise

    def display_processing_summary(self):
        """Display a summary of the processing results."""
        print(f"\n📊 Processing Summary")
        print(f"=" * 30)
        print(f"Files Processed: {self.stats['files_processed']}")
        print(f"Sheets Created: {self.stats['sheets_created']}")
        print(f"Matches Found: {self.stats['matches_found']}")
        print(f"Entries Added: {self.stats['entries_added']}")
        if self.stats['errors_encountered'] > 0:
            print(f"⚠️  Errors: {self.stats['errors_encountered']}")

        info_id(f"Processing summary: {self.stats}", self.request_id)

    def process_bom_loadsheet(self, source_file=None, max_matches=5):
        """Main processing method with enhanced error handling."""
        try:
            print(f"\n🔄 BOM Loadsheet Processing")
            print(f"=" * 40)

            # Get source file
            if not source_file:
                source_file = self.prompt_for_source_file()
            else:
                is_valid, message = self.validate_source_file(source_file)
                if not is_valid:
                    raise ValueError(f"Invalid source file: {message}")

            # Generate target file path
            target_file = self.generate_target_file_path(source_file)

            # Copy BOM sheet
            self.copy_bom_sheet_with_validation(source_file, target_file)

            # Match items and update sheet
            self.match_items_and_update_sheet(target_file, max_matches)

            # Display summary
            self.display_processing_summary()

            print(f"\n✅ BOM Loadsheet Processing Completed Successfully!")
            print(f"📄 Output file: {os.path.basename(target_file)}")
            info_id("BOM loadsheet processing completed successfully", self.request_id)

            return True

        except Exception as e:
            error_id(f"BOM loadsheet processing failed: {str(e)}", self.request_id, exc_info=True)
            print(f"\n❌ Processing failed: {str(e)}")
            return False


def main():
    """
    Main function to process BOM loadsheets.
    Uses the new PostgreSQL framework with enhanced error handling and user experience.
    """
    print("\n🎯 Starting BOM Loadsheet Processing")
    print("=" * 50)

    processor = None
    try:
        # Initialize the PostgreSQL processor
        processor = PostgreSQLBOMLoadsheetProcessor()

        # Ask for configuration
        max_matches = 5
        try:
            user_max = input(f"🔢 Maximum matches to process (default: {max_matches}): ").strip()
            if user_max and user_max.isdigit():
                max_matches = int(user_max)
                print(f"   📊 Set maximum matches to: {max_matches}")
        except:
            print(f"   📊 Using default maximum matches: {max_matches}")

        # Process the BOM loadsheet
        success = processor.process_bom_loadsheet(max_matches=max_matches)

        if success:
            print("\n✅ BOM Loadsheet Processing Completed Successfully!")
            print("=" * 50)
        else:
            print("\n⚠️  BOM Loadsheet Processing Completed with Issues")
            print("=" * 50)

    except KeyboardInterrupt:
        print("\n🛑 Processing interrupted by user")
        if processor:
            error_id("Processing interrupted by user", processor.request_id)
    except Exception as e:
        print(f"\n❌ Processing failed: {str(e)}")
        if processor:
            error_id(f"Processing failed: {str(e)}", processor.request_id, exc_info=True)
    finally:
        # Close logger
        try:
            close_initializer_logger()
        except:
            pass


if __name__ == "__main__":
    main()