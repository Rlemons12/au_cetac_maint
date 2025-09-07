import os
import sys
import logging
import pandas as pd
from typing import Optional, Dict, Any
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook, Workbook
from sqlalchemy.orm import Session
from modules.emtacdb.emtacdb_fts import (
    SiteLocation, Area, EquipmentGroup, Model, AssetNumber, Location,
    Subassembly, ComponentAssembly, AssemblyView, Position, Part, Image,
    PartsPositionImageAssociation
)
from modules.configuration.config import DATABASE_URL, DATABASE_DIR, LOAD_FOLDER_INTAKE, LOAD_FOLDER_OUTPUT, \
    LOAD_FOLDER_REFERENCE
from modules.configuration.config_env import DatabaseConfig
from modules.configuration.log_config import logger

##############################################################################
# Folder Constants (from config)
##############################################################################
# Files to load:
# Intake files are in LOAD_FOLDER_INTAKE.
# Output (master mapper, load_parts_with_images.xlsx) go to LOAD_FOLDER_OUTPUT.
# Reference photo list is in LOAD_FOLDER_REFERENCE.

##############################################################################
# Database Setup
##############################################################################
db_config = DatabaseConfig()
session = db_config.get_main_session()


##############################################################################
# Utility Functions
##############################################################################
def get_or_create(
        session: Session,
        model,
        defaults: Optional[Dict[str, Any]] = None,
        **kwargs
):
    """
    Check if an instance of `model` exists matching kwargs.
    If not, create it using the combined keyword arguments and `defaults`.
    Returns the instance (existing or newly created).
    """
    instance = session.query(model).filter_by(**kwargs).first()
    if instance:
        logger.debug(f"Found existing {model.__name__} entry: {kwargs}")
        return instance
    else:
        params = {**kwargs}
        if defaults:
            params.update(defaults)
        instance = model(**params)
        session.add(instance)
        session.commit()
        logger.info(f"Created new {model.__name__} entry: {params}")
        return instance


def get_cell_value(val):
    """
    Helper function to check if a cell value is empty.
    If the value is NaN (or a blank string), return None.
    Otherwise, return the value.
    """
    if pd.isna(val):
        return None
    if isinstance(val, str) and val.strip() == "":
        return None
    return val


##############################################################################
# Global MASTER_MAPPING
##############################################################################
# Maps the intake file's "id" (as string) to the DB-generated Position.id
MASTER_MAPPING = {}


##############################################################################
# Master Mapper Update Function
##############################################################################
def update_master_mapper(new_intake_df, source_filename):
    """
    Update (or create) the master_mapper.xlsx workbook with new intake data.
    Also adds a column for the DB-generated Position.id from MASTER_MAPPING.
    """
    master_mapper_path = os.path.join(LOAD_FOLDER_OUTPUT, "master_mapper.xlsx")
    new_df = new_intake_df.copy()
    new_df["source_file"] = source_filename

    def map_position_id(row):
        intake_id = get_cell_value(row['id'])
        if intake_id:
            key = str(intake_id).strip()
            return MASTER_MAPPING.get(key, None)
        return None

    new_df["position_id"] = new_df.apply(map_position_id, axis=1)

    if os.path.exists(master_mapper_path):
        try:
            master_df = pd.read_excel(master_mapper_path, sheet_name="Master")
            logger.debug("Existing master mapper loaded successfully.")
        except Exception as e:
            logger.error(f"Error reading existing master mapper: {e}")
            master_df = pd.DataFrame()
        combined_df = pd.concat([master_df, new_df], ignore_index=True)
        logger.info("Combined new intake data with existing master mapper data.")
    else:
        combined_df = new_df
        logger.info("No existing master mapper found; using new intake data.")

    try:
        with pd.ExcelWriter(master_mapper_path, engine='openpyxl') as writer:
            combined_df.to_excel(writer, sheet_name="Master", index=False)
        logger.info(f"Master mapper updated at {master_mapper_path}")
    except Exception as e:
        logger.error(f"Error saving master mapper: {e}")


##############################################################################
# Intake File Processing (for 'Intake' sheet)
##############################################################################
def load_intake_file(file_path: str, session: Session) -> None:
    """
    Reads the 'Intake' sheet, creates/updates DB records, and updates MASTER_MAPPING.
    Note: We have removed 'assembly_view' and 'assembly_view_descriptoion' from the required columns.
    """
    logger.info(f"Attempting to load Excel file: {file_path}")
    try:
        intake_df = pd.read_excel(file_path, sheet_name="Intake")
        intake_df.columns = intake_df.columns.str.strip()
        logger.info(f"Successfully loaded Excel file: {file_path}")
    except Exception as e:
        logger.error(f"Failed to load Excel file: {file_path}. Error: {e}")
        return

    # Updated expected columns; assembly_view fields are optional now.
    expected_columns = [
        'id', 'title', 'room_number', 'site_area', 'area', 'area_description',
        'equipment_group', 'equipment_group_Description', 'model', 'model_description',
        'asset_number', 'asset_number_description', 'locations', 'locations_description',
        'subassembly', 'subassembly_decription', 'component_assembly', 'component_assembly_decription'
    ]
    missing_columns = [col for col in expected_columns if col not in intake_df.columns]
    if missing_columns:
        logger.error(f"Missing expected columns: {missing_columns}. Skipping file {file_path}.")
        return

    for index, row in intake_df.iterrows():
        logger.info(f"Processing row {index + 1} in file {file_path}")
        area_val = get_cell_value(row['area'])
        equipment_group_val = get_cell_value(row['equipment_group'])
        model_val = get_cell_value(row['model'])
        asset_number_val = get_cell_value(row['asset_number'])
        locations_val = get_cell_value(row['locations'])
        if not all([area_val, equipment_group_val, model_val, asset_number_val, locations_val]):
            logger.warning(f"Row {index + 1}: Required fields empty. Skipping row.")
            continue
        try:
            area = get_or_create(
                session, Area, name=area_val,
                defaults={'description': get_cell_value(row['area_description'])}
            )
            equipment_group = get_or_create(
                session, EquipmentGroup, name=equipment_group_val, area_id=area.id,
                defaults={'description': get_cell_value(row['equipment_group_Description'])}
            )
            model = get_or_create(
                session, Model, name=model_val, equipment_group_id=equipment_group.id,
                defaults={'description': get_cell_value(row['model_description'])}
            )
            asset_number = get_or_create(
                session, AssetNumber, number=asset_number_val, model_id=model.id,
                defaults={'description': get_cell_value(row['asset_number_description'])}
            )
            location = get_or_create(
                session, Location, name=locations_val, model_id=model.id,
                defaults={'description': get_cell_value(row['locations_description'])}
            )

            subassembly_val = get_cell_value(row['subassembly'])
            if subassembly_val:
                subassembly = get_or_create(
                    session, Subassembly, name=subassembly_val, location_id=location.id,
                    defaults={'description': get_cell_value(row['subassembly_decription'])}
                )
            else:
                subassembly = None

            component_assembly_val = get_cell_value(row['component_assembly'])
            if component_assembly_val:
                if subassembly is None:
                    logger.warning(f"Row {index + 1}: Component Assembly but no Subassembly. Skipping it.")
                    component_assembly = None
                else:
                    component_assembly = get_or_create(
                        session, ComponentAssembly, name=component_assembly_val, subassembly_id=subassembly.id,
                        defaults={'description': get_cell_value(row['component_assembly_decription'])}
                    )
            else:
                component_assembly = None

            site_area_val = get_cell_value(row['site_area'])
            title_val = get_cell_value(row['title'])
            room_number_val = get_cell_value(row['room_number'])
            if site_area_val or title_val or room_number_val:
                site_location = get_or_create(
                    session, SiteLocation,
                    site_area=site_area_val, title=title_val, room_number=room_number_val,
                    defaults={}
                )
            else:
                site_location = None

            site_location_id = site_location.id if site_location else None

            # Check if Position already exists; if not, create new Position.
            position = session.query(Position).filter_by(
                area_id=area.id,
                equipment_group_id=equipment_group.id,
                model_id=model.id,
                asset_number_id=asset_number.id,
                location_id=location.id,
                subassembly_id=(subassembly.id if subassembly else None),
                component_assembly_id=(component_assembly.id if component_assembly else None),
                site_location_id=site_location_id
            ).first()

            if position:
                logger.debug(f"Row {index + 1}: Position already exists in file {file_path}")
                intake_id = get_cell_value(row['id'])
                if intake_id:
                    MASTER_MAPPING[str(intake_id).strip()] = position.id
            else:
                new_position = Position(
                    area_id=area.id,
                    equipment_group_id=equipment_group.id,
                    model_id=model.id,
                    asset_number_id=asset_number.id,
                    location_id=location.id,
                    subassembly_id=(subassembly.id if subassembly else None),
                    component_assembly_id=(component_assembly.id if component_assembly else None),
                    site_location_id=site_location_id
                )
                session.add(new_position)
                session.commit()
                logger.info(f"Row {index + 1}: Added new Position in file {file_path}")
                intake_id = get_cell_value(row['id'])
                if intake_id:
                    MASTER_MAPPING[str(intake_id).strip()] = new_position.id

        except Exception as e:
            logger.error(f"Row {index + 1}: Error processing file {file_path}: {e}", exc_info=True)

    source_filename = os.path.basename(file_path)
    update_master_mapper(intake_df, source_filename)


##############################################################################
# Parts Sheet Processing (with image matching)
##############################################################################
def process_row_for_parts(sheet, item_number, position_id, photo, description, manufacturer_description):
    """
    Append a row to the parts output sheet if a photo is provided.
    The output row now consists of:
      - Item Number (with 'A' prefix)
      - Position ID
      - Image (with 'A' prefix)
      - Description
      - Full Description (concatenation of description and manufacturer_description)
    """
    if photo:
        full_description = f"{description}, {manufacturer_description}" if manufacturer_description else description
        final_item_number = item_number if str(item_number).startswith("A") else f"A{item_number}"
        final_photo = photo if str(photo).startswith("A") else f"A{photo}"
        sheet.append([final_item_number, position_id, final_photo, description, full_description])
        logger.info(
            f"Added entry: Item Number: {final_item_number}, Position ID: {position_id}, "
            f"Image: {final_photo}, Description: {full_description}"
        )


##############################################################################
# Parts Sheet Processing & Output Sheet Creation
##############################################################################
def load_parts_file(file_path: str, session: Session) -> None:
    """
    Process the 'parts' sheet from the specified Excel file, match part numbers with images,
    and output a new workbook 'load_parts_with_images.xlsx' (in the output folder) with the results.
    Then this output sheet will be used by load_part_position_image_associations() to insert into the DB.

    For matching, the script strips an 'A' from the item number for matching,
    but for output it re-adds 'A' to both the item number and photo strings.
    """
    logger.info(f"Attempting to load Parts sheet from {file_path}")
    try:
        parts_df = pd.read_excel(file_path, sheet_name="parts")
        parts_df.columns = parts_df.columns.str.strip()
        logger.info(f"Successfully loaded the 'parts' sheet from {file_path}")
    except Exception as e:
        logger.error(f"Failed to load the 'parts' sheet from {file_path}. Error: {e}")
        return

    expected_columns = [
        "Manufacturer", "Mfg Part Number", "Asset Number", "Item Number",
        "Description", "Long Description", "Inventory Location", "Qty On Hand",
        "Category", "id"
    ]
    missing_columns = [col for col in expected_columns if col not in parts_df.columns]
    if missing_columns:
        logger.error(f"Missing expected columns in parts sheet: {missing_columns}. Skipping file {file_path}.")
        return

    try:
        intake_df = pd.read_excel(file_path, sheet_name="Intake")
        intake_df.columns = intake_df.columns.str.strip()
        logger.info("Loaded Intake sheet for position mapping.")
    except Exception as e:
        logger.error(f"Failed to load the 'Intake' sheet from {file_path}. Error: {e}")
        intake_df = None

    # Build fallback dictionary from asset -> { area, equipment_group, model }
    intake_mapping = {}
    if intake_df is not None:
        for _, intake_row in intake_df.iterrows():
            asset = get_cell_value(intake_row['asset_number'])
            if asset:
                cleaned_asset = asset.strip()
                intake_mapping[cleaned_asset] = {
                    "id": get_cell_value(intake_row['id']),
                    "area": get_cell_value(intake_row['area']),
                    "equipment_group": get_cell_value(intake_row['equipment_group']),
                    "model": get_cell_value(intake_row['model'])
                }

    # Prepare the target workbook in the output folder
    target_file_name = "load_parts_with_images.xlsx"
    target_file_path = os.path.join(LOAD_FOLDER_OUTPUT, target_file_name)
    if os.path.exists(target_file_path):
        wb_target = load_workbook(target_file_path)
        if "part_position_image" in wb_target.sheetnames:
            part_sheet = wb_target["part_position_image"]
            logger.info("Using existing 'part_position_image' sheet in target workbook.")
        else:
            part_sheet = wb_target.create_sheet("part_position_image")
            part_sheet.append(["Item Number", "Position ID", "Image", "Description", "Full Description"])
            logger.info("Created new sheet 'part_position_image' in target workbook.")
    else:
        wb_target = Workbook()
        default_sheet = wb_target.active
        wb_target.remove(default_sheet)
        part_sheet = wb_target.create_sheet("part_position_image")
        part_sheet.append(["Item Number", "Position ID", "Image", "Description", "Full Description"])
        logger.info(f"Created new target workbook: {target_file_path} with sheet 'part_position_image'.")

    # Load the photo list workbook from the reference folder
    part_list_image_path = os.path.join(LOAD_FOLDER_REFERENCE, "part_list_image.xlsx")
    if not os.path.exists(part_list_image_path):
        logger.error(f"Photo list workbook not found at {part_list_image_path}. Cannot match images.")
        return
    try:
        wb_photo = load_workbook(part_list_image_path)
        if "photo_list" not in wb_photo.sheetnames:
            logger.error(f"'photo_list' sheet not found in {part_list_image_path}.")
            return
        photo_list_sheet = wb_photo["photo_list"]
        logger.info("Successfully loaded 'photo_list' sheet.")
    except Exception as e:
        logger.error(f"Failed to load photo list workbook: {e}")
        return

    def insert_row(item_num_final, position_id, photo, desc, manufacturer_desc):
        """
        Append a row to 'part_position_image' if 'photo' is present,
        ensuring both the item number and photo have an 'A' prefix.
        """
        if photo:
            final_item_num = item_num_final if item_num_final.startswith("A") else f"A{item_num_final}"
            final_photo = photo if photo.startswith("A") else f"A{photo}"
            full_desc = f"{desc}, {manufacturer_desc}" if manufacturer_desc else desc
            part_sheet.append([final_item_num, position_id, final_photo, desc, full_desc])
            logger.info(
                f"Added entry: Item Number: {final_item_num}, Position ID: {position_id}, "
                f"Image: {final_photo}, Description: {full_desc}"
            )

    with ThreadPoolExecutor() as executor:
        for index, row in parts_df.iterrows():
            raw_item_number = get_cell_value(row["Item Number"])
            if not raw_item_number:
                logger.warning(f"Row {index + 1} in parts sheet: Missing Item Number; skipping row.")
                continue
            raw_item_number = str(raw_item_number).strip()
            # For matching, remove the 'A' prefix if present
            item_number_noA = raw_item_number[1:] if raw_item_number.startswith("A") else raw_item_number
            # For final output, ensure the item number has the 'A' prefix
            item_num_final = raw_item_number if raw_item_number.startswith("A") else f"A{raw_item_number}"
            # Use first 6 characters of the non-'A' version for matching with photo list
            part_prefix = item_number_noA[:6]

            parts_asset = get_cell_value(row["Asset Number"])
            if parts_asset:
                if parts_asset.startswith("AU-"):
                    cleaned_asset = parts_asset[3:]
                else:
                    cleaned_asset = parts_asset
                if "-" in cleaned_asset:
                    cleaned_asset = cleaned_asset.split("-")[0]
            else:
                cleaned_asset = ""

            parts_id_value = get_cell_value(row["id"])
            if parts_id_value:
                intake_key = str(parts_id_value).strip()
                if intake_key in MASTER_MAPPING:
                    position_id = MASTER_MAPPING[intake_key]
                    logger.debug(f"Row {index + 1}: Using Position ID from MASTER_MAPPING: {position_id}")
                else:
                    logger.warning(f"Row {index + 1}: Intake id {intake_key} not in MASTER_MAPPING. Using fallback.")
                    position_id = parts_id_value
            else:
                mapping = intake_mapping.get(cleaned_asset)
                if mapping:
                    area_obj = session.query(Area).filter_by(name=mapping['area']).first()
                    eq_group_obj = session.query(EquipmentGroup).filter_by(name=mapping['equipment_group']).first()
                    model_obj = session.query(Model).filter_by(name=mapping['model']).first()
                    if area_obj and eq_group_obj and model_obj:
                        existing_position = session.query(Position).filter_by(
                            area_id=area_obj.id,
                            equipment_group_id=eq_group_obj.id,
                            model_id=model_obj.id,
                            asset_number_id=None,
                            location_id=None,
                            subassembly_id=None,
                            component_assembly_id=None,
                            assembly_view_id=None,
                            site_location_id=None
                        ).first()
                        if existing_position:
                            position_id = existing_position.id
                            logger.debug(
                                f"Row {index + 1}: Found existing generic Position with ID {position_id} as fallback.")
                        else:
                            new_position = Position(
                                area_id=area_obj.id,
                                equipment_group_id=eq_group_obj.id,
                                model_id=model_obj.id,
                                asset_number_id=None,
                                location_id=None,
                                subassembly_id=None,
                                component_assembly_id=None,
                                assembly_view_id=None,
                                site_location_id=None
                            )
                            session.add(new_position)
                            session.commit()
                            position_id = new_position.id
                            logger.info(
                                f"Row {index + 1}: Created new generic Position with ID {position_id} as fallback.")
                    else:
                        logger.warning(f"Row {index + 1}: Could not create new Position. Missing required objects.")
                        position_id = ""
                else:
                    logger.warning(
                        f"Row {index + 1}: No matching intake row for asset '{cleaned_asset}'. Using empty Position ID.")
                    position_id = ""

            # Match with photo list workbook
            for photo_row in photo_list_sheet.iter_rows(min_row=2, values_only=True):
                photo_part_number = str(photo_row[0])
                photo_prefix = photo_part_number[:6]
                if part_prefix == photo_prefix:
                    logger.info(
                        f"Row {index + 1}: Match found: item_number_noA '{item_number_noA}' matches photo list part '{photo_part_number}'")
                    photo_a = photo_row[1]
                    photo_b = photo_row[2]
                    photo_c = photo_row[3]
                    desc_a = photo_row[4]
                    desc_b = photo_row[5]
                    desc_c = photo_row[6]
                    manufacturer_desc = photo_row[7]

                    # Submit tasks for each photo
                    executor.submit(insert_row, item_num_final, position_id, photo_a, desc_a, manufacturer_desc)
                    executor.submit(insert_row, item_num_final, position_id, photo_b, desc_b, manufacturer_desc)
                    executor.submit(insert_row, item_num_final, position_id, photo_c, desc_c, manufacturer_desc)

    wb_target.save(target_file_path)
    logger.info(f"Parts with images saved to {target_file_path}")


##############################################################################
# Load Part-Position-Image Associations
##############################################################################
def load_part_position_image_associations(session: Session) -> None:
    """
    Reads the 'part_position_image' sheet from 'load_parts_with_images.xlsx'
    and updates/inserts data into the PartsPositionImageAssociation table.
    """
    file_path = os.path.join(LOAD_FOLDER_OUTPUT, "load_parts_with_images.xlsx")
    if not os.path.exists(file_path):
        logger.error(f"Parts with images file not found at {file_path}. Aborting.")
        return

    logger.info(f"Loading workbook from {file_path} for PartsPositionImageAssociation processing.")
    try:
        wb = load_workbook(file_path)
        if "part_position_image" not in wb.sheetnames:
            logger.error("Sheet 'part_position_image' not found in load_parts_with_images.xlsx. Aborting.")
            return
        sheet = wb["part_position_image"]
        logger.info("Successfully loaded 'part_position_image' sheet.")
    except Exception as e:
        logger.error(f"Error loading workbook: {e}")
        return

    row_count = 0
    assoc_created_or_updated = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_count += 1
        # Columns: Item Number, Position ID, Image, Description, Full Description
        part_number = row[0]
        position_id = row[1]
        image_title = row[2]

        logger.debug(
            f"Row {row_count}: Part Number='{part_number}', Position ID='{position_id}', Image Title='{image_title}'"
        )
        if not part_number or not position_id:
            logger.warning(f"Row {row_count}: Missing Part Number or Position ID; skipping.")
            continue

        final_part_number = part_number if str(part_number).startswith("A") else f"A{part_number}"
        logger.debug(f"Row {row_count}: Using final_part_number='{final_part_number}'.")

        final_image_title = None
        if image_title:
            final_image_title = image_title if str(image_title).startswith("A") else f"A{image_title}"
            logger.debug(f"Row {row_count}: Using final_image_title='{final_image_title}'.")
        else:
            logger.info(f"Row {row_count}: No image title provided; proceeding without image.")

        part_obj = session.query(Part).filter_by(part_number=final_part_number).first()
        if not part_obj:
            logger.warning(f"Row {row_count}: Part with part_number '{final_part_number}' not found; skipping row.")
            continue

        image_obj = None
        if final_image_title:
            image_obj = session.query(Image).filter_by(title=final_image_title).first()
            if not image_obj:
                logger.warning(f"Row {row_count}: No Image found for title '{final_image_title}'; image_id=None.")
        try:
            existing_assoc = session.query(PartsPositionImageAssociation).filter_by(
                part_id=part_obj.id,
                position_id=position_id,
                image_id=image_obj.id if image_obj else None
            ).first()

            if existing_assoc:
                logger.info(
                    f"Row {row_count}: Association already exists for Part '{final_part_number}', Position '{position_id}', Image '{final_image_title}'.")
            else:
                new_assoc = PartsPositionImageAssociation(
                    part_id=part_obj.id,
                    position_id=position_id,
                    image_id=image_obj.id if image_obj else None
                )
                session.add(new_assoc)
                session.commit()
                assoc_created_or_updated += 1
                logger.info(
                    f"Row {row_count}: Created new association for Part '{final_part_number}', Position '{position_id}', Image '{final_image_title}'.")
        except Exception as e:
            session.rollback()
            logger.error(
                f"Row {row_count}: Failed to process association for Part '{final_part_number}', Position '{position_id}', Image '{final_image_title}'. Error: {e}")

    logger.info(
        f"Finished processing associations. Processed {row_count} rows; created/updated {assoc_created_or_updated} associations."
    )


##############################################################################
# File Processing Options
##############################################################################
def process_single_file(session: Session) -> None:
    file_name = input("Enter the file name (with extension) to load: ")
    file_path = os.path.join(LOAD_FOLDER_INTAKE, file_name)
    if os.path.exists(file_path) and file_name.startswith("load_intake_"):
        # 1) Load intake
        load_intake_file(file_path, session)
        # 2) Create the "part_position_image" sheet from parts data
        load_parts_file(file_path, session)
        # 3) Now read the "part_position_image" sheet and insert associations into the DB
        load_part_position_image_associations(session)
    else:
        logger.error(f"File {file_name} does not exist or does not follow naming convention 'load_intake_'.")


def process_multiple_files(session: Session) -> None:
    for file_name in os.listdir(LOAD_FOLDER_INTAKE):
        if file_name.startswith("load_intake_") and file_name.endswith(".xlsx"):
            file_path = os.path.join(LOAD_FOLDER_INTAKE, file_name)
            load_intake_file(file_path, session)
            load_parts_file(file_path, session)
    load_part_position_image_associations(session)


##############################################################################
# Main Entry Point
##############################################################################
def main():
    logger.info("Select processing mode:")
    logger.info("1: Process a single file")
    logger.info("2: Process all files in the folder")
    mode = input("Enter your choice (1 or 2): ").strip()
    if mode == '1':
        process_single_file(session)
    elif mode == '2':
        process_multiple_files(session)
    else:
        logger.error("Invalid processing mode selected.")
    logger.info("Data loading process completed.")


if __name__ == "__main__":
    main()
