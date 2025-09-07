from flask import Blueprint, render_template, request, redirect, url_for, flash
import os
import logging
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from modules.configuration.config import DB_LOADSHEET_BOMS, DATABASE_PATH_IMAGES_FOLDER  # Added DATABASE_PATH_IMAGES_FOLDER
from modules.configuration.config_env import DatabaseConfig
from modules.emtacdb.emtacdb_fts import Part, PartsPositionImageAssociation
from modules.emtacdb.utlity.main_database.database import create_position, add_image_to_db
from sqlalchemy.orm.exc import NoResultFound
import shutil  # Added to handle file copying
from utilities.auth_utils import login_required
from modules.configuration.log_config import logger
bill_of_materials_bp = Blueprint('bill_of_materials_bp', __name__, template_folder='templates')



db_config = DatabaseConfig()

def allowed_file(filename):
    logger.info(f"Checking if file {filename} is allowed.")
    allowed = '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx'}
    logger.info(f"File {filename} allowed: {allowed}")
    return allowed

def prompt_for_target_file(source_file):
    logging.info(f"Generating target file name based on source file: {source_file}")
    suffix = os.path.basename(source_file).replace("bom_for_", "").replace(".xlsx", "")

    if not os.path.exists(DB_LOADSHEET_BOMS):
        os.makedirs(DB_LOADSHEET_BOMS)
        logging.info(f"Directory {DB_LOADSHEET_BOMS} created.")

    target_file_name = f"load_bom_for_{suffix}.xlsx"
    file_path = os.path.join(DB_LOADSHEET_BOMS, target_file_name)

    if os.path.exists(file_path):
        logging.info(f"Target file already exists: {file_path}")
    else:
        logging.info(f"Creating new target file: {file_path}")

    return file_path

def copy_bom_sheet_to_target(source_path, target_path):
    logging.info(f"Copying BOM sheet from {source_path} to {target_path}.")
    wb_source = load_workbook(source_path)
    if "BOM" in wb_source.sheetnames:
        bom_sheet = wb_source["BOM"]

        if os.path.exists(target_path):
            wb_target = load_workbook(target_path)
        else:
            wb_target = Workbook()
            default_sheet = wb_target.active
            wb_target.remove(default_sheet)

        suffix = os.path.basename(target_path).replace("load_bom_for_", "").replace(".xlsx", "")
        new_bom_sheet_name = f"bom_{suffix}"
        bom_target_sheet = wb_target.create_sheet(new_bom_sheet_name)

        for row in bom_sheet.iter_rows(values_only=True):
            bom_target_sheet.append(row)

        part_position_image_sheet = wb_target.create_sheet("part_position_image")
        part_position_image_sheet.append(["part", "position", "image", "description"])

        wb_target.save(target_path)
        logging.info(f"'BOM' sheet copied to '{new_bom_sheet_name}' in {target_path}.")
    else:
        logging.error(f"The source workbook does not contain a sheet named 'BOM'. Source: {source_path}")

@bill_of_materials_bp.route('/bill_of_materials', methods=['GET', 'POST'])
@login_required
def bill_of_materials():
    logger.info("Accessed bill_of_materials route.")
    if request.method == 'POST':
        logger.info("Received POST request.")

        # Retrieve the image path from the form
        image_path = request.form.get('image_path')
        logger.info(f"Image path received: {image_path}")
        if not image_path:
            flash('Image path is required')
            logging.error("Image path is required but not provided.")
            return redirect(request.url)

        # Retrieve data from the form for Position creation
        area_id = request.form.get('area')
        equipment_group_id = request.form.get('equipment_group')
        model_id = request.form.get('model')
        asset_number_id = request.form.get('asset_number')
        location_id = request.form.get('location')
        site_location_id = request.form.get('site_location')
        logger.info(f"Received Position data: Area ID={area_id}, Equipment Group ID={equipment_group_id}, "
                      f"Model ID={model_id}, Asset Number ID={asset_number_id}, Location ID={location_id}, "
                      f"Site Location ID={site_location_id}")

        session = db_config.get_main_session()

        # Create the Position using the create_position function
        position_id = create_position(area_id=area_id, equipment_group_id=equipment_group_id, model_id=model_id,
                                      asset_number_id=asset_number_id, location_id=location_id,
                                      site_location_id=site_location_id, session=session)

        if position_id is None:
            flash('Error creating position')
            logging.error("Error creating position.")
            return redirect(request.url)

        flash('Position created successfully!')
        logging.info(f"Position created successfully with ID: {position_id}")

        # Check if the POST request has the file part
        if 'file' not in request.files:
            flash('No file part')
            logging.error("No file part in the POST request.")
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('No selected file')
            logging.error("No selected file in the POST request.")
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(DB_LOADSHEET_BOMS, filename)
            file.save(file_path)
            flash('File successfully uploaded')
            logging.info(f"File successfully uploaded: {file_path}")

            try:
                process_bom_loadsheet(file_path, image_path, position_id)  # Pass the position_id here
                flash('File successfully processed')
                logging.info(f"File successfully processed: {file_path}")
            except Exception as e:
                logging.error(f"Error processing file {file_path}: {e}")
                flash(f"An error occurred while processing the file: {e}")

        return redirect(url_for('bill_of_materials_bp.bill_of_materials'))

    return render_template('bill_of_materials/bill_of_materials.html')

def process_bom_loadsheet(source_file, image_path, position_id):
    logging.info(f"Starting BOM loadsheet process for file: {source_file}")

    # Generate the target file path based on the uploaded file
    target_path = prompt_for_target_file(source_file)

    # Copy the BOM sheet to the new sheet in the target workbook
    copy_bom_sheet_to_target(source_file, target_path)

    # Match items using the part_list_image sheet and update the part_position_image sheet
    match_items_in_part_list_image(target_path, image_path, position_id)

    # Process remaining items by searching the image folder
    process_remaining_items_in_image_folder(target_path, image_path, position_id)

    logging.info("BOM loadsheet process completed.")

def match_items_in_part_list_image(target_path, image_path, position_id):
    """
    Match items from a BOM sheet with a part list image sheet and update the target workbook with matched images.

    Args:
        target_path (str): Path to the target Excel workbook containing the BOM sheet.
        image_path (str): Path where the images are stored.
        position_id (int): ID of the position associated with the parts.

    Side Effects:
        Updates the 'part_position_image' sheet in the target workbook with matched images and saves the workbook.
    """
    try:
        logging.info(f"Matching items with part_list_image sheet in {target_path}.")

        wb_target = load_workbook(target_path)
        bom_sheet_name = [sheet for sheet in wb_target.sheetnames if sheet.startswith("bom_")][0]
        bom_sheet = wb_target[bom_sheet_name]
        part_position_image_sheet = wb_target["part_position_image"]

        part_list_image_path = os.path.join(DB_LOADSHEET_BOMS, "part_list_image.xlsx")
        wb_part_list = load_workbook(part_list_image_path)
        photo_list_sheet = wb_part_list["photo_list"]

        # Process all rows in the BOM sheet
        for row_idx, row in enumerate(bom_sheet.iter_rows(min_row=2, values_only=True), start=2):
            item_number = str(row[3]).strip()  # Assuming "Item Number" is the fourth column in BOM sheet
            logger.info(f"Processing item_number: '{item_number}' at BOM row {row_idx}")

            # Remove the leading "A" and get the first 6 characters
            part_number_prefix = item_number[1:7].upper()  # Strip the leading "A" and get the next 6 characters
            match_found = False

            for photo_row in photo_list_sheet.iter_rows(min_row=2, values_only=True):
                photo_part_number_prefix = str(photo_row[0])[:6].strip().upper()  # First 6 characters of Part #
                logger.info(
                    f"Comparing BOM prefix '{part_number_prefix}' with Part # prefix '{photo_part_number_prefix}'")

                if part_number_prefix == photo_part_number_prefix:
                    logger.info(
                        f"Match found in part_list_image for item number prefix: {part_number_prefix} at BOM row {row_idx}")

                    # Process photos from the Excel file
                    photo_a = photo_row[1]
                    desc_a = photo_row[4]  # Corresponding "Desc A"
                    photo_b = photo_row[2]
                    desc_b = photo_row[5]  # Corresponding "Desc B"
                    photo_c = photo_row[3]
                    desc_c = photo_row[6]  # Corresponding "Desc C"
                    manufacturer_description = photo_row[7]  # "Manufacturer Description"

                    # Prefix photo names with the correct part number
                    prefixed_photo_a = f"{item_number[:1]}{photo_a}" if photo_a else None
                    prefixed_photo_b = f"{item_number[:1]}{photo_b}" if photo_b else None
                    prefixed_photo_c = f"{item_number[:1]}{photo_c}" if photo_c else None

                    # Log and process each photo
                    if prefixed_photo_a:
                        logger.info(f"Passing prefixed photo name to process_part_position_image: {prefixed_photo_a}")
                        process_part_position_image(item_number, position_id, prefixed_photo_a, image_path)
                    if prefixed_photo_b:
                        logger.info(f"Passing prefixed photo name to process_part_position_image: {prefixed_photo_b}")
                        process_part_position_image(item_number, position_id, prefixed_photo_b, image_path)
                    if prefixed_photo_c:
                        logger.info(f"Passing prefixed photo name to process_part_position_image: {prefixed_photo_c}")
                        process_part_position_image(item_number, position_id, prefixed_photo_c, image_path)

                    match_found = True
                    break  # Stop searching once a match is found

            if not match_found:
                logger.info(f"No match found in part_list_image for item number: {item_number} at BOM row {row_idx}")

        # Save the final workbook after processing all rows
        wb_target.save(target_path)
        logging.info(f"part_position_image sheet updated with part_list_image matches in {target_path}.")

    except FileNotFoundError as fnf_error:
        logging.error(f"File not found: {fnf_error}")
    except KeyError as key_error:
        logging.error(f"Key error: {key_error}")
    except Exception as e:
        logging.error(f"An error occurred: {e}")

def process_row(part_position_image_sheet, item_number, photo, description, manufacturer_description, position_id):
    logger.info(
        f"Preparing to process row: Item Number: {item_number}, Photo: {photo}, Description: {description}, Manufacturer Description: {manufacturer_description}, Position ID: {position_id}")

    if photo:
        full_description = f"{description}, {manufacturer_description}" if manufacturer_description else description
        part_position_image_sheet.append([item_number, position_id, photo, full_description])
        logging.info(
            f"Added entry: Item Number: {item_number}, Position ID: {position_id}, Photo: {photo}, Description: {full_description}")
    else:
        logging.warning(f"No photo provided for Item Number: {item_number}. Row not added.")

def find_image_in_subfolders(image_title, base_path):
    logger.info(f"Searching for image '{image_title}' in '{base_path}' and its subfolders.")
    """
    Search for the image in the given base_path and its subfolders.
    Returns the full file path if found, otherwise returns None.
    """
    for root, dirs, files in os.walk(base_path):
        logger.info(f"Searching in directory: {root}")  # Log each directory being searched
        for file in files:
            # Get the file name without extension
            file_name_without_ext = os.path.splitext(file)[0]
            # Check if the file name matches the image_title
            if file_name_without_ext == image_title:
                full_path = os.path.join(root, file)
                logging.info(f"Image found: {full_path}")
                return full_path
    logging.warning(f"Image '{image_title}' not found in '{base_path}' or its subfolders.")
    return None

def process_remaining_items_in_image_folder(target_path, image_path, position_id):
    logging.info(f"Processing all BOM items by searching the image folder from the beginning in {target_path}.")
    wb_target = load_workbook(target_path)
    bom_sheet_name = [sheet for sheet in wb_target.sheetnames if sheet.startswith("bom_")][0]
    bom_sheet = wb_target[bom_sheet_name]
    part_position_image_sheet = wb_target["part_position_image"]

    for row_idx, row in enumerate(bom_sheet.iter_rows(min_row=2, values_only=True), start=2):
        item_number = str(row[3])  # Assuming "Item Number" is the fourth column in BOM sheet
        logger.info(f"Processing BOM item_number: '{item_number}' at row {row_idx}")

        if item_number:
            first_seven_chars = item_number[:7]  # Extract first 7 characters from the item number
            logger.info(f"Attempting to match BOM item '{item_number}' with files in image folder using '{first_seven_chars}'.")

            # Search for a matching image in the folder
            image_file_path = find_image_in_subfolders(first_seven_chars, image_path)

            if image_file_path:
                logging.info(f"Match found for BOM item '{item_number}' at row {row_idx} with image file '{image_file_path}'.")
                process_part_position_image(item_number, position_id, first_seven_chars, image_path)
            else:
                logging.warning(f"No matching image found for BOM item '{item_number}' at row {row_idx}. Creating entry with no image.")
                create_part_position_entry_no_image(item_number, position_id)

    wb_target.save(target_path)
    logging.info(f"part_position_image sheet updated with image folder matches in {target_path}.")

def create_part_position_entry_no_image(part_number, position_id):
    """
    Create a record in part_position_image for parts without corresponding images.
    """
    session = db_config.get_main_session()
    try:
        part = session.query(Part).filter(Part.part_number == part_number[:7]).one()
        logging.info(f"Part found with ID: {part.id} for part number: {part_number}")

        # Create an entry in the part_position_image table without an image
        part_position_image_association = PartsPositionImageAssociation(
            part_id=part.id,
            position_id=position_id,
            image_id=None  # No image available
        )

        session.add(part_position_image_association)
        session.commit()
        logging.info(f"Entry created in part_position_image with Part ID: {part.id}, Position ID: {position_id}, without image.")
    except NoResultFound:
        logging.error(f"No part found with part number: {part_number}.")
        session.rollback()

def process_part_position_image(part_number, position_id, image_title, base_image_path):
    logging.info(f"Starting to process part position image for Part Number: {part_number}, Position ID: {position_id}, Image Title: {image_title}")

    # Locate the image file in the base path or its subfolders
    logger.info(f"Attempting to locate image '{image_title}' in '{base_image_path}' or its subfolders.")
    image_file_path = find_image_in_subfolders(image_title, base_image_path)

    if not image_file_path:
        logging.error(f"Image '{image_title}' not found in {base_image_path} or its subfolders. Aborting process for this image.")
        return

    # **New Code Block**: Save the image to `DATABASE_PATH_IMAGES_FOLDER`
    filename = os.path.basename(image_file_path)
    saved_image_path = os.path.join(DATABASE_PATH_IMAGES_FOLDER, filename)
    shutil.copy(image_file_path, saved_image_path)
    logging.info(f"Image '{filename}' copied to '{saved_image_path}'.")

    # Use the file name (without extension) as the title
    file_name_without_ext = os.path.splitext(os.path.basename(saved_image_path))[0]
    logger.info(f"Using file name without extension as the image title: {file_name_without_ext}")

    # **Updated**: Add the image to the database with the new path
    image_id = add_image_to_db(title=file_name_without_ext, file_path=saved_image_path, position_id=position_id)

    if image_id:
        logging.info(f"Image '{file_name_without_ext}' processed and added to the database with ID {image_id}.")
        create_part_position_image_association(file_name_without_ext, position_id, image_id, db_config.get_main_session())
    else:
        logging.error(f"Failed to process image '{file_name_without_ext}'. It was not added to the database.")

    logging.info(f"Finished processing part position image for Part Number: {part_number}, Position ID: {position_id}, Image Title: {file_name_without_ext}")

def create_part_position_image_association(image_title, position_id, image_id, session):
    logging.info(f"Creating part_position_image association for image title: {image_title}, position ID: {position_id}, image ID: {image_id}")
    # Extract the part number from the image title (first 7 characters)
    part_number = image_title[:7]
    logger.info(f"Extracted part number from image title: {part_number}")

    # Look for the part in the Part table
    try:
        part = session.query(Part).filter(Part.part_number == part_number).one()
        logging.info(f"Part found with ID: {part.id} for part number: {part_number}")

        # Create a new entry in the part_position_image table
        part_position_image_association = PartsPositionImageAssociation(
            part_id=part.id,
            position_id=position_id,
            image_id=image_id
        )

        session.add(part_position_image_association)
        session.commit()
        logging.info(f"Entry created in part_position_image with Part ID: {part.id}, Position ID: {position_id}, Image ID: {image_id}")

    except NoResultFound:
        logging.error(f"No part found with part number: {part_number}. Unable to create part_position_image association.")
        session.rollback()

